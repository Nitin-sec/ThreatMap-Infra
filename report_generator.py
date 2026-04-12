"""
report_generator.py — ThreatMap Infra Professional VAPT Report
Three sheets: Cover Page → Observations → Annexure

Design: White/light background — print-safe, professional.
AI: observation_name, detailed_observation, impacted_module, risk_impact
    are now properly displayed (were previously missing from report).
Storage: user chooses output folder; filename includes target + timestamp.
"""

import os
import glob
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Print-safe palette (white background) ────────────────────────────────────
WHITE      = "FFFFFF"
OFF_WHITE  = "F8F9FA"
LIGHT_GRAY = "F0F0F0"
MID_GRAY   = "DDDDDD"
DARK_GRAY  = "555555"
NEAR_BLACK = "1A1A1A"
ACCENT_RED = "C0392B"
ACCENT_NAV = "2C3E50"
MUTED      = "777777"
BORDER_C   = "CCCCCC"

SEV_BG = {
    "Critical": "FDECEA",
    "High":     "FEF3E2",
    "Medium":   "FFFDE7",
    "Low":      "F1F8E9",
    "Info":     "E8F4FD",
}
SEV_FG = {
    "Critical": "B71C1C",
    "High":     "BF360C",
    "Medium":   "E65100",
    "Low":      "1B5E20",
    "Info":     "0D47A1",
}
SEV_BADGE = {
    "Critical": "C0392B",
    "High":     "E67E22",
    "Medium":   "F39C12",
    "Low":      "27AE60",
    "Info":     "2980B9",
}
SEV_ORDER = ["Critical", "High", "Medium", "Low", "Info"]

DEFAULTS = {
    "host": "", "port": "", "service": "", "severity": "Info",
    "priority_rank": 5, "cvss_score": 0.0, "actively_exploited": 0,
    "observation_name": "", "detailed_observation": "",
    "impacted_module": "Network Service", "risk_impact": "",
    "risk_summary": "", "business_impact": "", "remediation": "",
    "false_positive_likelihood": "Low", "attack_scenario": "",
    "triage_method": "rule_based", "ai_enhanced": 0,
}


# ── Style helpers ─────────────────────────────────────────────────────────────

def _fill(c):
    return PatternFill("solid", fgColor=c)

def _font(color=NEAR_BLACK, bold=False, size=10, name="Arial", italic=False):
    return Font(color=color, bold=bold, size=size, name=name, italic=italic)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _border(c=BORDER_C):
    s = Side(style="thin", color=c)
    return Border(left=s, right=s, top=s, bottom=s)

def _bg(ws, rows, cols, color=WHITE):
    for r in range(1, rows + 1):
        for c in range(1, cols + 1):
            ws.cell(row=r, column=c).fill = _fill(color)

def _to_dict(row):
    d = dict(DEFAULTS)
    d.update({k: v for k, v in dict(row).items() if v is not None})
    if d["severity"] not in SEV_ORDER:
        d["severity"] = "Info"
    return d


# ── Sheet 1: Cover Page ───────────────────────────────────────────────────────

def _cover(wb, records, meta):
    ws = wb.active
    ws.title = "Cover Page"
    ws.sheet_view.showGridLines = False
    _bg(ws, 60, 14, WHITE)

    for i, w in enumerate([2,22,4,30,4,16,4,14,4,2,2,2,2,2], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Red top bar
    ws.row_dimensions[1].height = 8
    for c in range(1, 15):
        ws.cell(row=1, column=c).fill = _fill(ACCENT_RED)

    # Brand name
    ws.row_dimensions[3].height = 50
    c = ws.cell(row=3, column=2, value="THREATMAP")
    c.font = Font(name="Arial", bold=True, size=38, color=ACCENT_RED)
    c.alignment = _align("left", "center")
    ws.merge_cells("B3:H3")

    ws.row_dimensions[4].height = 20
    c = ws.cell(row=4, column=2,
                value="INFRA  —  Vulnerability Assessment & Penetration Testing Report")
    c.font = Font(name="Arial", size=12, color=MUTED)
    c.alignment = _align("left", "center")
    ws.merge_cells("B4:H4")

    # Divider
    ws.row_dimensions[5].height = 2
    for c in range(2, 9):
        ws.cell(row=5, column=c).fill = _fill(ACCENT_RED)

    # Metadata
    target = meta.get("target") or (records[0]["host"] if records else "—")
    date   = meta.get("date") or datetime.now().strftime("%d %B %Y")
    mode   = (meta.get("mode") or "balanced").title()

    for row, label, value in [
        (7,  "Target",         target),
        (8,  "Report Date",    date),
        (9,  "Scan Mode",      mode),
        (10, "Classification", "CONFIDENTIAL — Authorised Recipients Only"),
        (11, "Prepared By",    "ThreatMap Infra v1.0 — Automated VAPT Engine"),
    ]:
        ws.row_dimensions[row].height = 20
        lc = ws.cell(row=row, column=2, value=label)
        lc.font = Font(name="Arial", bold=True, size=9, color=MUTED)
        vc = ws.cell(row=row, column=4, value=value)
        vc.font = Font(name="Arial", size=10, color=NEAR_BLACK)
        vc.alignment = _align("left", "center")
        ws.merge_cells(f"D{row}:I{row}")

    # Divider
    ws.row_dimensions[13].height = 2
    for c in range(2, 9):
        ws.cell(row=13, column=c).fill = _fill(MID_GRAY)

    # Severity counts
    counts = {s: 0 for s in SEV_ORDER}
    for r in records:
        counts[r["severity"]] = counts.get(r["severity"], 0) + 1

    ws.row_dimensions[14].height = 10
    ws.row_dimensions[15].height = 50
    ws.row_dimensions[16].height = 18
    col = 2
    for sev in SEV_ORDER:
        cnt = counts.get(sev, 0)
        bg  = SEV_BADGE[sev]
        # Count
        c = ws.cell(row=15, column=col, value=str(cnt))
        c.font = Font(name="Arial", bold=True, size=28, color=bg)
        c.alignment = _align("center", "center")
        ws.merge_cells(f"{get_column_letter(col)}15:{get_column_letter(col+1)}15")
        # Label
        c = ws.cell(row=16, column=col, value=sev.upper())
        c.font = Font(name="Arial", bold=True, size=8, color=WHITE)
        c.fill = _fill(bg)
        c.alignment = _align("center", "center")
        ws.merge_cells(f"{get_column_letter(col)}16:{get_column_letter(col+1)}16")
        col += 2

    # Scope
    ws.row_dimensions[19].height = 14
    c = ws.cell(row=19, column=2, value="SCOPE & METHODOLOGY")
    c.font = Font(name="Arial", bold=True, size=9, color=ACCENT_NAV)

    scope = (
        f"This assessment was conducted against {target} using automated "
        "reconnaissance and vulnerability scanning tools including Nmap, Nikto, "
        "Gobuster, SSLScan, WhatWeb, and Nuclei. Findings are classified per "
        "CVSS v3.1 and prioritised by exploitability and business impact. "
        "AI-assisted analysis enriches findings where configured."
    )
    ws.row_dimensions[20].height = 52
    c = ws.cell(row=20, column=2, value=scope)
    c.font = Font(name="Arial", size=9, color=MUTED)
    c.fill = _fill(OFF_WHITE)
    c.alignment = _align("left", "top", wrap=True)
    ws.merge_cells("B20:I20")

    # Footer
    ws.row_dimensions[55].height = 2
    for c in range(1, 15):
        ws.cell(row=55, column=c).fill = _fill(ACCENT_RED)
    ws.row_dimensions[56].height = 18
    c = ws.cell(row=56, column=2,
                value="CONFIDENTIAL — Contains sensitive security information. "
                      "Not for distribution beyond authorised recipients.")
    c.font = Font(name="Arial", size=8, color=WHITE)
    c.fill = _fill(ACCENT_RED)
    c.alignment = _align("center")
    ws.merge_cells("B56:I56")


# ── Sheet 2: Observations ─────────────────────────────────────────────────────

def _observations(wb, records):
    ws = wb.create_sheet("Observations")
    ws.sheet_view.showGridLines = False
    _bg(ws, 500, 12, WHITE)

    for col, w in {"A":2,"B":7,"C":32,"D":40,"E":22,"F":14,"G":40,"H":44,"I":2}.items():
        ws.column_dimensions[col].width = w

    # Title
    ws.row_dimensions[1].height = 8
    for c in range(1, 10):
        ws.cell(row=1, column=c).fill = _fill(ACCENT_RED)
    ws.row_dimensions[2].height = 30
    c = ws.cell(row=2, column=2, value="OBSERVATIONS")
    c.font = Font(name="Arial", bold=True, size=18, color=ACCENT_RED)
    c.alignment = _align("left", "center")

    # Summary
    counts = {s: 0 for s in SEV_ORDER}
    for r in records:
        counts[r["severity"]] = counts.get(r["severity"], 0) + 1
    parts = [f"{counts[s]} {s}" for s in SEV_ORDER if counts.get(s, 0) > 0]
    ws.row_dimensions[3].height = 16
    c = ws.cell(row=3, column=2,
                value=f"Total: {len(records)}   ·   " + "   ·   ".join(parts))
    c.font = Font(name="Arial", size=9, color=MUTED)
    c.alignment = _align("left", "center")
    ws.merge_cells("B3:H3")

    ws.row_dimensions[4].height = 3
    for c in range(2, 9):
        ws.cell(row=4, column=c).fill = _fill(MID_GRAY)

    # Headers
    HDRS = [(2,"S. No","center"),(3,"Observation Name","left"),
            (4,"Detailed Observation","left"),(5,"Impacted Module","center"),
            (6,"Severity","center"),(7,"Risk / Impact","left"),
            (8,"Recommendation","left")]
    ws.row_dimensions[5].height = 22
    for col, hdr, align in HDRS:
        c = ws.cell(row=5, column=col, value=hdr)
        c.fill = _fill(ACCENT_NAV)
        c.font = Font(name="Arial", bold=True, size=9, color=WHITE)
        c.alignment = _align(align, "center")
        c.border = _border(WHITE)

    ws.freeze_panes = "B6"

    sorted_recs = sorted(
        records,
        key=lambda r: (int(r["priority_rank"]), -(float(r["cvss_score"])))
    )

    for i, rec in enumerate(sorted_recs, 1):
        row    = 5 + i
        ws.row_dimensions[row].height = 80
        sev    = rec["severity"]
        sbg    = SEV_BG.get(sev, OFF_WHITE)
        sfg    = SEV_FG.get(sev, DARK_GRAY)
        badge  = SEV_BADGE.get(sev, DARK_GRAY)
        is_hot = sev in ("Critical", "High")
        row_bg = OFF_WHITE if i % 2 == 0 else WHITE

        # Prefer AI fields, fall back to rule-based
        obs    = rec["observation_name"] or f"Exposed {(rec['service'] or 'Unknown').upper()} Service"
        detail = rec["detailed_observation"] or rec["risk_summary"] or \
                 f"Port {rec['port']}/TCP ({(rec['service'] or '').upper()}) is publicly accessible."
        mod    = rec["impacted_module"] or "Network Service"
        risk   = rec["risk_impact"] or rec["business_impact"] or "Impact not assessed."
        fix    = rec["remediation"] or "Refer to vendor security guidance."
        ai_tag = " ✦" if rec.get("ai_enhanced") else ""

        def _cell(col, val, bg=row_bg, fg=DARK_GRAY, bold=False, halign="left", wrap=True):
            c = ws.cell(row=row, column=col, value=val)
            c.fill      = _fill(bg)
            c.font      = Font(name="Arial", bold=bold, size=9, color=fg)
            c.alignment = _align(halign, "top", wrap)
            c.border    = _border()
            return c

        _cell(2, i,            bg=sbg if is_hot else row_bg,
              fg=sfg if is_hot else MUTED, bold=True, halign="center", wrap=False)
        _cell(3, obs + ai_tag, bg=sbg if is_hot else row_bg,
              fg=sfg if is_hot else NEAR_BLACK, bold=True)
        _cell(4, detail)
        _cell(5, mod, halign="center")

        # Severity badge (white text on colour)
        sc = ws.cell(row=row, column=6, value=sev)
        sc.fill      = _fill(badge)
        sc.font      = Font(name="Arial", bold=True, size=9, color=WHITE)
        sc.alignment = _align("center", "center")
        sc.border    = _border()

        _cell(7, risk)
        _cell(8, fix)

    # Footer note
    last = 5 + len(sorted_recs) + 2
    ws.row_dimensions[last].height = 14
    c = ws.cell(last, 2, value="✦ = AI-enhanced   · Sorted by severity and priority")
    c.font = Font(name="Arial", size=8, color=MUTED, italic=True)
    ws.merge_cells(f"B{last}:H{last}")


# ── Sheet 3: Annexure ─────────────────────────────────────────────────────────

def _annexure(wb, meta):
    ws = wb.create_sheet("Annexure")
    ws.sheet_view.showGridLines = False
    _bg(ws, 5000, 12, WHITE)

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 110
    ws.column_dimensions["D"].width = 2

    ws.row_dimensions[1].height = 8
    for c in range(1, 12):
        ws.cell(row=1, column=c).fill = _fill(ACCENT_RED)

    ws.row_dimensions[2].height = 30
    c = ws.cell(row=2, column=2, value="ANNEXURE — SCAN EVIDENCE")
    c.font = Font(name="Arial", bold=True, size=18, color=ACCENT_RED)
    c.alignment = _align("left", "center")

    ws.row_dimensions[3].height = 16
    c = ws.cell(row=3, column=2,
                value="Raw scan tool outputs. All data stored locally — no server uploads.")
    c.font = Font(name="Arial", size=9, color=MUTED)
    c.alignment = _align("left", "center")

    row     = 5
    reports = "reports"
    target  = meta.get("target", "target")

    SECTIONS = [
        ("NMAP — PORT SCAN",     f"nmap_{target}.xml"),
        ("WHOIS — REGISTRATION", f"whois_{target}.txt"),
        ("DNS — ENUMERATION",    f"dig_{target}.txt"),
        ("NIKTO — WEB SCAN",     f"nikto_{target}.txt"),
        ("GOBUSTER — DIRS",      f"gobuster_{target}.txt"),
        ("SSLSCAN — TLS",        f"sslscan_{target}.txt"),
        ("NUCLEI — CVE SCAN",    f"nuclei_{target}.txt"),
        ("CURL — HTTP HEADERS",  f"curl_headers_{target}.txt"),
        ("WHATWEB — TECH STACK", f"whatweb_{target}.json"),
        ("AUTHORIZATION LOG",    "authorization_log.txt"),
        ("SCAN LOG",             "scan.log"),
    ]
    for ef in glob.glob(f"{reports}/evidence_*.json"):
        SECTIONS.append((f"HTTP EVIDENCE — {Path(ef).name}", Path(ef).name))

    written = 0
    for title, filename in SECTIONS:
        fpath = os.path.join(reports, filename)
        if not os.path.isfile(fpath):
            matches = glob.glob(os.path.join(reports, filename.replace(target, "*")))
            if matches:
                fpath = matches[0]
        if not os.path.isfile(fpath):
            continue

        try:
            content = Path(fpath).read_text(encoding="utf-8", errors="replace").strip()
        except Exception:
            continue
        if not content:
            continue

        written += 1

        # Section header
        ws.row_dimensions[row].height = 20
        c = ws.cell(row=row, column=2, value=title)
        c.fill = _fill(ACCENT_NAV)
        c.font = Font(name="Arial", bold=True, size=10, color=WHITE)
        c.alignment = _align("left", "center")
        c.border = _border(WHITE)
        ws.merge_cells(f"B{row}:C{row}")
        row += 1

        # Source path
        ws.row_dimensions[row].height = 12
        c = ws.cell(row=row, column=2, value=f"Source: {fpath}")
        c.fill = _fill(LIGHT_GRAY)
        c.font = Font(name="Arial", size=7, color=MUTED)
        c.alignment = _align("left", "center")
        ws.merge_cells(f"B{row}:C{row}")
        row += 1

        lines = content.splitlines()
        if len(lines) > 300:
            lines = lines[:300]
            lines.append(f"[truncated — see {fpath}]")

        for line in lines:
            ws.row_dimensions[row].height = 12
            c = ws.cell(row=row, column=2, value=line)
            c.fill = _fill(OFF_WHITE if row % 2 == 0 else WHITE)
            c.font = Font(name="Courier New", size=8, color=DARK_GRAY)
            c.alignment = _align("left", "center")
            ws.merge_cells(f"B{row}:C{row}")
            row += 1

        ws.row_dimensions[row].height = 6
        row += 1

    if written == 0:
        ws.row_dimensions[row].height = 20
        c = ws.cell(row=row, column=2, value="No scan logs found — run a scan first.")
        c.font = Font(name="Arial", size=10, color=MUTED)
        c.alignment = _align("left", "center")


# ── Public API ────────────────────────────────────────────────────────────────

def generate_excel(db, output_dir: str = "reports") -> str | None:
    """
    Build the professional three-sheet VAPT report.
    Saves to output_dir (chosen by user at scan time).
    Filename: ThreatMap_<target>_<YYYYMMDD_HHMM>.xlsx
    """
    raw = db.get_all_triage()
    if not raw:
        print("[!] No triage findings to export.")
        return None

    # Convert sqlite3.Row → plain dict (fixes AttributeError on .get())
    records = [_to_dict(r) for r in raw]

    meta = {}
    try:
        with db._conn() as conn:
            row = conn.execute(
                "SELECT target, scan_mode, started_at, completed_at "
                "FROM scans ORDER BY id DESC LIMIT 1"
            ).fetchone()
            if row:
                meta = {
                    "target": row["target"],
                    "mode":   row["scan_mode"],
                    "date":   (row["completed_at"] or row["started_at"] or "")[:10],
                }
    except Exception:
        pass

    target_slug = (
        (meta.get("target") or "scan")
        .replace(".", "_").replace("/", "_").replace(":", "")
    )
    ts       = datetime.now().strftime("%Y%m%d_%H%M")
    filename = f"ThreatMap_{target_slug}_{ts}.xlsx"
    path     = os.path.join(output_dir, filename)

    wb = openpyxl.Workbook()
    _cover(wb, records, meta)
    _observations(wb, records)
    _annexure(wb, meta)

    wb["Cover Page"].sheet_properties.tabColor   = "C0392B"
    wb["Observations"].sheet_properties.tabColor = "2980B9"
    wb["Annexure"].sheet_properties.tabColor     = "7F8C8D"
    wb.active = wb["Cover Page"]

    os.makedirs(output_dir, exist_ok=True)
    wb.save(path)

    total  = len(records)
    hosts  = len({r["host"] for r in records})
    ai_n   = sum(1 for r in records if r.get("ai_enhanced"))
    counts = {s: 0 for s in SEV_ORDER}
    for r in records:
        counts[r["severity"]] = counts.get(r["severity"], 0) + 1

    print(f"[✔] Report → {path}")
    print(f"    {total} findings · {hosts} host(s) · {ai_n} AI-enhanced")
    print(f"    " + "  ".join(f"{s}:{counts[s]}" for s in SEV_ORDER if counts.get(s)))

    return path
